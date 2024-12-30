import datetime
import logging
import string
from tokenize import Comment
from fastapi import APIRouter, Depends, HTTPException, Header, File, Request, UploadFile
from typing import List, Dict
from app.services import pipefy_service
from app.core.security import get_current_user, decrypt_token
from app.models.user import User
from app.db.mongodb import MongoDB
from io import BytesIO
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from fastapi import APIRouter, Depends, HTTPException, Header, File, Request, UploadFile, Body
from pydantic import BaseModel
from bson import ObjectId

router = APIRouter()
logger = logging.getLogger(__name__)

# Dicionário para armazenar temporariamente os campos de cada usuário
user_fields_cache = {}

# Modelo Pydantic para os campos selecionados
class SelectedFieldsModel(BaseModel):
    selected_fields: List[str]

class TemplateGenerationModel(BaseModel):
    selected_fields: List[str]
    selected_user: str 

class PipeCreate(BaseModel):
    name: str
    pipeId: str

class PipeUpdate(BaseModel):
    name: str
    pipeId: str

class PipeInDB(BaseModel):
    id: str
    name: str
    pipeId: str
    user_id: str

async def get_pipefy_token(current_user: User = Depends(get_current_user)):
    logger.info(f"Retrieving Pipefy token for user: {current_user.email}")
    user = await MongoDB.database.users.find_one({"email": current_user.email})
    if not user or "pipefy_token" not in user:
        logger.warning(f"Pipefy token not found for user: {current_user.email}")
        raise HTTPException(status_code=400, detail="Pipefy token not found. Please save your Pipefy token first.")
    token = decrypt_token(user["pipefy_token"])
    logger.info(f"Successfully retrieved and decrypted token for user: {current_user.email}")
    logger.debug(f"Token: {token[:4]}...{token[-4:]}")
    return token

@router.post("/get_phases")
async def get_phases(pipe_id: str, current_user: User = Depends(get_current_user), authorization: str = Header(None)):
    logger.info(f"Received request for pipe_id: {pipe_id}")
    logger.info(f"Authorization header: {authorization}")
    
    if pipe_id.startswith("http"):
        pipe_id = pipe_id.split("/")[-1]
    
    try:
        api_token = await get_pipefy_token(current_user)
        logger.info(f"Calling Pipefy API for pipe_id: {pipe_id}")
        
        # Buscar fases e membros do pipe
        phases = pipefy_service.get_pipe_phases(pipe_id, api_token)
        pipe_members = pipefy_service.get_pipe_members(pipe_id, api_token)
        
        # Armazenar no cache
        user_fields_cache[current_user.email] = {
            "pipe_id": pipe_id,
            "phases": phases,
            "pipe_members": pipe_members
        }
        
        return {"phases": phases}
    except Exception as e:
        logger.error(f"Error fetching pipe phases: {str(e)}", exc_info=True)
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/get_fields")
async def get_fields(phase_id: str, current_user: User = Depends(get_current_user)):
    try:
        api_token = await get_pipefy_token(current_user)
        fields = pipefy_service.get_phase_fields(phase_id, api_token)
        
        # Recuperar membros do pipe do cache
        user_data = user_fields_cache.get(current_user.email, {})
        pipe_members = user_data.get('pipe_members', [])
        
        # Armazenar os campos e o phase_id para este usuário
        user_fields_cache[current_user.email].update({
            "phase_id": phase_id, 
            "fields": fields
        })
        
        return {"fields": fields}
    except Exception as e:
        logger.error(f"Error fetching fields: {str(e)}", exc_info=True)
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/prepare_fields_selection")
async def prepare_fields_selection(
    selected_fields: List[str], 
    current_user: User = Depends(get_current_user)
):
    try:
        user_data = user_fields_cache.get(current_user.email)
        
        if not user_data:
            raise HTTPException(status_code=400, detail="Please fetch fields first")
        
        all_fields = user_data.get('fields', [])
        pipe_members = user_data.get('pipe_members', [])
        
        # Validar se os campos selecionados existem
        field_map = {field['label']: field for field in all_fields}
        for field_label in selected_fields:
            if field_label not in field_map:
                raise HTTPException(status_code=400, detail=f"Campo {field_label} não encontrado")
        
        # Identificar campos assignee_select entre os campos selecionados
        assignee_fields = [
            field for field in all_fields 
            if field['label'] in selected_fields and field['type'] == 'assignee_select'
        ]
        
        # Preparar informações de membros para campos assignee
        assignee_options = []
        if assignee_fields:
            assignee_options = [
                {
                    "field_label": field['label'],
                    "members": [
                        {
                            "id": member['user']['id'],
                            "name": member['user']['name'],
                            "email": member['user']['email']
                        } for member in pipe_members
                    ]
                } for field in assignee_fields
            ]
        
        # Armazenar informações para próximas etapas
        user_data['selected_fields'] = selected_fields
        user_data['assignee_fields'] = assignee_fields
        
        return {
            "selected_fields": selected_fields,
            "assignee_options": assignee_options
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao preparar seleção de campos: {str(e)}", exc_info=True)
        raise HTTPException(status_code=400, detail=str(e))

@router.api_route("/get_pipe_members", methods=["POST", "OPTIONS"])
async def get_pipe_members(request: Request, pipe_id: str = None, current_user: User = Depends(get_current_user)):
    if request.method == "OPTIONS":
        return JSONResponse(content={"message": "OK"}, status_code=200)
    
    if not pipe_id:
        raise HTTPException(status_code=400, detail="pipe_id is required")
    
    try:
        # Extrair o ID do pipe da URL, se necessário
        if pipe_id.startswith("http"):
            pipe_id = pipe_id.split("/")[-1]
        
        api_token = await get_pipefy_token(current_user)
        logger.info(f"Fetching members for pipe_id: {pipe_id}")
        members = pipefy_service.get_pipe_members(pipe_id, api_token)
        
        # Log da resposta para depuração
        logger.debug(f"Members fetched: {members}")
        
        return {"members": members}
    except Exception as e:
        logger.error(f"Error fetching pipe members: {str(e)}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"Error fetching pipe members: {str(e)}")

@router.post("/generate_xlsx_template")
async def generate_xlsx_template(data: TemplateGenerationModel, current_user: User = Depends(get_current_user)):
    try:
        user_data = user_fields_cache.get(current_user.email)
        
        if not user_data:
            raise HTTPException(status_code=400, detail="Please prepare field selection first")
        
        all_fields = user_data.get('fields', [])
        
        logger.info(f"Received selected fields: {data.selected_fields}")
        logger.info(f"Received selected user: {data.selected_user}")
        logger.info(f"All fields: {all_fields}")
        
        # Filtrar campos selecionados, incluindo campos assignee
        selected_field_details = [
            field for field in all_fields 
            if field['id'] in data.selected_fields
        ]
        
        logger.info(f"Selected field details: {selected_field_details}")
        
        wb = Workbook()
        ws = wb.active
        
        # Cabeçalho visível do XLSX (apenas labels)
        visible_headers = ["ID do card"] + [field['label'] for field in selected_field_details]
        ws.append(visible_headers)
        
        # Cabeçalho oculto com IDs dos campos
        hidden_headers = ["card_id"] + [field['id'] for field in selected_field_details]
        ws.append(hidden_headers)
        
        # Ocultar a linha com os IDs
        ws.row_dimensions[2].hidden = True
        
        # Adicionar o responsável selecionado como valor padrão para campos assignee
        for row in ws.iter_rows(min_row=3, max_row=3):
            for cell, field in zip(row[1:], selected_field_details):
                if field['type'] == 'assignee_select':
                    cell.value = data.selected_user
        
        # Ajustar a largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # Salvar o arquivo
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(iter([buffer.getvalue()]),
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment;filename=update_template.xlsx"})
    
    except Exception as e:
        logger.error(f"Erro ao gerar template XLSX: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erro ao gerar template: {str(e)}")

@router.post("/update_cards_from_xlsx")
async def update_cards_from_xlsx(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    try:
        user_data = user_fields_cache.get(current_user.email)
        
        if not user_data:
            raise HTTPException(status_code=400, detail="Please prepare field selection first")
        
        contents = await file.read()
        wb = load_workbook(filename=BytesIO(contents), data_only=True)
        ws = wb.active
        
        api_token = await get_pipefy_token(current_user)
        
        results = []
        
        visible_headers = [cell.value for cell in ws[1]]
        field_ids = [cell.value for cell in ws[2]]  # Hidden row with field IDs
        
        for row_index, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not row[0]:  # Skip if no card ID
                continue
                
            card_id = str(row[0])
            field_updates = {}
            
            for i, value in enumerate(row[1:], start=1):
                if value is not None:
                    field_id = field_ids[i]
                    field_updates[field_id] = str(value)
            
            if not field_updates:
                logger.warning(f"No updates for card {card_id}")
                continue
            
            try:
                success, message = pipefy_service.update_card_fields(card_id, field_updates, api_token)
                results.append({"card_id": card_id, "success": success, "message": message})
            except Exception as e:
                logger.error(f"Error updating card {card_id}: {str(e)}", exc_info=True)
                results.append({"card_id": card_id, "success": False, "message": str(e)})
        
        if not results:
            logger.warning("No cards were updated")
        else:
            logger.info(f"Update results: {results}")
        
        return {"results": results}
    except Exception as e:
        logger.error(f"Error updating cards from XLSX: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error updating cards: {str(e)}")
    
@router.post("/pipes", response_model=PipeInDB)
async def create_pipe(pipe: PipeCreate, current_user: User = Depends(get_current_user)):
    try:
        new_pipe = PipeInDB(
            id=str(ObjectId()),
            name=pipe.name,
            pipeId=pipe.pipeId,
            user_id=str(current_user.id)
        )
        result = await MongoDB.database.pipes.insert_one(new_pipe.dict())
        logger.info(f"Created new pipe with id: {result.inserted_id}")
        return new_pipe
    except Exception as e:
        logger.error(f"Error creating pipe: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error creating pipe: {str(e)}")

@router.put("/pipes/{pipe_id}", response_model=PipeInDB)
async def update_pipe(pipe_id: str, pipe: PipeUpdate, current_user: User = Depends(get_current_user)):
    try:
        updated_pipe = await MongoDB.database.pipes.find_one_and_update(
            {"_id": ObjectId(pipe_id), "user_id": str(current_user.id)},
            {"$set": pipe.dict()},
            return_document=True
        )
        if not updated_pipe:
            logger.warning(f"Pipe not found: {pipe_id}")
            raise HTTPException(status_code=404, detail="Pipe not found")
        logger.info(f"Updated pipe: {pipe_id}")
        return PipeInDB(**{k: v for k, v in updated_pipe.items() if k != '_id'}, id=str(updated_pipe["_id"]))
    except Exception as e:
        logger.error(f"Error updating pipe: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error updating pipe: {str(e)}")

@router.delete("/pipes/{pipe_id}", response_model=dict)
async def delete_pipe(pipe_id: str, current_user: User = Depends(get_current_user)):
    result = await MongoDB.database.pipes.delete_one({"_id": ObjectId(pipe_id), "user_id": str(current_user.id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pipe not found")
    return {"message": "Pipe deleted successfully"}

@router.get("/pipes", response_model=List[PipeInDB])
async def get_pipes(current_user: User = Depends(get_current_user)):
    logger.info(f"Fetching pipes for user: {current_user.id}")
    pipes = await MongoDB.database.pipes.find({"user_id": str(current_user.id)}).to_list(None)
    logger.info(f"Found {len(pipes)} pipes")
    result = [
        PipeInDB(
            id=str(pipe["_id"]),
            name=pipe["name"],
            pipeId=pipe["pipeId"],
            user_id=pipe["user_id"]
        )
        for pipe in pipes
    ]
    logger.info(f"Returning {len(result)} pipes")
    return result

@router.post("/save-template")
async def save_template(
    template_data: dict,
    current_user: User = Depends(get_current_user)
):
    template = {
        "name": template_data["name"],
        "pipe_id": template_data["pipe_id"],
        "phase_id": template_data["phase_id"],
        "fields": template_data["fields"],
        "user_id": str(current_user.id)
    }
    result = await MongoDB.database.templates.insert_one(template)
    return {"id": str(result.inserted_id), "message": "Template saved successfully"}

@router.get("/templates")
async def get_templates(current_user: User = Depends(get_current_user)):
    logger.info(f"Buscando templates para o usuário: {current_user.id}")
    
    try:
        templates = await MongoDB.database.templates.find({"user_id": str(current_user.id)}).to_list(None)
        logger.info(f"Templates encontrados: {len(templates)}")
        
        result = [
            {
                "id": str(template["_id"]),
                "name": template.get("name", "Template sem nome"),
                "pipe_id": template.get("pipe_id", ""),
                "phase_id": template.get("phase_id", ""),
                "fields": template.get("fields", [])
            }
            for template in templates
        ]
        
        logger.info(f"Templates processados: {result}")
        return result
    
    except Exception as e:
        logger.error(f"Erro ao buscar templates: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erro interno ao buscar templates: {str(e)}")


@router.delete("/templates/{template_id}")
async def delete_template(
    template_id: str,
    current_user: User = Depends(get_current_user)
):
    result = await MongoDB.database.templates.delete_one({"_id": ObjectId(template_id), "user_id": str(current_user.id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"message": "Template deleted successfully"}
